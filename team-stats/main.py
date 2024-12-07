"""
This code is intended to run on Cloud Functions.
This code can analyze basketball team statistics.
"""

# import
import pandas as pd
from google.cloud import storage
from io import BytesIO
from google.cloud import bigqueryß
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import pytz
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
import base64
from email.mime.text import MIMEText
import time
import pickle
import os
from google.cloud import storage
from datetime import timedelta


# 東京のタイムゾーンを設定
tokyo_tz = pytz.timezone('Asia/Tokyo')
# 現在の日時を取得し、東京時間に変換
tokyo_time = datetime.now(tokyo_tz)
# yyyymmdd形式でフォーマット
formatted_date = tokyo_time.strftime('%Y%m%d')


# function
def input_data(column_letter, data_list):
    start_row = 3  # 挿入を開始する行番号
    for i, value in enumerate(data_list, start=start_row):
        ws[f'{column_letter}{i}'] = value
def change_name(original_name):
    converted_name = ref_df[ref_df["名前"] == original_name]["name"].values[0]
    return converted_name
def change_color(search_name):
    for i in range(17):
        if ws['A'+str(i+3)].value == "Tokyo Z":
            # 塗る色を設定
            fill = PatternFill(start_color="93A8D7",  # 青色
                    end_color="93A8D7",
                    fill_type="solid")
            # 特定のセルに適用
            for col in col_list:
                ws[col+str(i+3)].fill = fill
        else:
            pass
def make_rank(target_name):
    for i in range(17):
        if ws['A'+str(i+3)].value == "Tokyo Z":
            num = str(i+3)
            # 特定のセルに適用
            for col in rank_list:
                ws[col+str(21)].value = f"=RANK({col}{num},{col}3:{col}19)"
# メール送信

def generate_signed_url(bucket_name, blob_name, expiration_minutes=60):
    """
    GCSオブジェクトの署名付きURLを生成
    """
    try:
        # GCSクライアントを初期化
        client = storage.Client()

        # バケットとオブジェクトを指定
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(blob_name)

        # 署名付きURLを生成
        url = blob.generate_signed_url(
            version="v4",
            expiration=timedelta(minutes=expiration_minutes),
            method="GET"
        )

        return url
    except Exception as e:
        print(f"An error occurred: {e}")
        raise



def get_gcs_public_url(bucket_name, file_name):
    """
    GCSファイルの公開URLを取得
    """
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(file_name)

    # ファイルを公開
    blob.make_public()

    # 公開URLを返す
    return blob.public_url

def sending_mail(bucket_name, blob_name):

    # GCSファイルの公開リンクを取得
    signed_url = generate_signed_url(bucket_name, blob_name)

    # Gmail APIの認証スコープ
    SCOPES = ['https://www.googleapis.com/auth/gmail.send']

    # OAuth 2.0の認証
    flow = InstalledAppFlow.from_client_secrets_file(
        '/root/xxxx.json', SCOPES
    )
    credentials = flow.run_local_server(port=0)

    # Gmail APIのクライアントを構築
    service = build('gmail', 'v1', credentials=credentials)

    # メール本文を作成
    message = MIMEText(f"""
        <html>
        <body>
        <p>お疲れ様です。</p>

        <p>先週までの最新断面の結果をお送りします！</p>

        <p><a href="{signed_url}" target="_blank">ダウンロードはこちらからお願いします。</a></p>
        <p>seto</p>
        </body>
        </html>
    """, "html")
    message['to'] = 'xxxxx'
    message['subject'] = "B3リーグスタッツ"

    # メールをエンコードして送信
    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    message = {'raw': raw}
    result = service.users().messages().send(userId='me', body=message).execute()
    print(f"Message ID: {result['id']}")

# main
# 書き込み用ファイルの準備
bucket_name = "b3_input_bucket"
file_path = "teamstats_edited.xlsx"
# GCS クライアントを初期化
client = storage.Client()
bucket = client.bucket(bucket_name)
blob = bucket.blob(file_path)
# GCS からファイルを読み込み
file_data = blob.download_as_bytes()
# メモリ上にファイルをロードして Excel を読み込む
wb = load_workbook(filename=BytesIO(file_data))
# シート名を指定して取得
sheet_name = 'Team Avg'
if sheet_name in wb.sheetnames:  # 指定したシートが存在するか確認
    ws = wb[sheet_name]
else:
    raise ValueError(f"指定されたシート名 '{sheet_name}' は存在しません。")

#ref
gcs_file_path = f"gs://{bucket_name}/{file_path}"
ref_df = pd.read_excel(gcs_file_path,sheet_name="Ref",names=["名前","name"])

# 書き込み内容の準備
client = bigquery.Client()
query = """
SELECT `クラブ名`, GP,W,L,`WIN%`,PTS, `3PM`, `3PA`, `3P%`, `2PM`, `2PA`, `2P%`, 
FTM, FTA, `FT%`, PF, OREB, DREB, REB, TOV, AST, STL, BLK,
FROM `xxxxxxxxxx` order by `WIN%` DESC
"""
df = client.query(query).to_dataframe()
df["クラブ名"] = df["クラブ名"].apply(change_name)
col_list = [
    "A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q",
    "R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF"
]
rank_list = [
    "E","F","G","H","I","J","K","L","M","N","O","P","Q",
    "R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF"
]

df["PTS"] = np.round(df["PTS"]/df["GP"],1)
df["3PM"] = np.round(df["3PM"]/df["GP"],1)
df["3PA"] = np.round(df["3PA"]/df["GP"],1)
df["2PM"] = np.round(df["2PM"]/df["GP"],1)
df["2PA"] = np.round(df["2PA"]/df["GP"],1)
df["FTM"] = np.round(df["FTM"]/df["GP"],1)
df["FTA"] = np.round(df["FTA"]/df["GP"],1)
df["PF"] = np.round(df["PF"]/df["GP"],1)
df["OREB"] = np.round(df["OREB"]/df["GP"],1)
df["DREB"] = np.round(df["DREB"]/df["GP"],1)
df["REB"] = np.round(df["REB"]/df["GP"],1)
df["TOV"] = np.round(df["TOV"]/df["GP"],1)
df["AST"] = np.round(df["AST"]/df["GP"],1)
df["STL"] = np.round(df["STL"]/df["GP"],1)
df["BLK"] = np.round(df["BLK"]/df["GP"],1)

# クラブ名
input_data("A", df["クラブ名"].to_list())
# GP
input_data("B", df["GP"].to_list())
# W
input_data("C", df["W"].to_list())
# L
input_data("D", df["L"].to_list())
# WIN%
input_data("E", [x * 0.01 for x in df["WIN%"].to_list()])
# PTS
input_data("F", df["PTS"].to_list())
# 3PM
input_data("J", df["3PM"].to_list())
# 3PA
input_data("K", df["3PA"].to_list())
# 2PM
input_data("M", df["2PM"].to_list())
# 2PA
input_data("N", df["2PA"].to_list())
# FTM
input_data("P", df["FTM"].to_list())
# FTA
input_data("Q", df["FTA"].to_list())
# PF
input_data("S", df["PF"].to_list())
# OREB
input_data("T", df["OREB"].to_list())
# DREB
input_data("U", df["DREB"].to_list())
# REB
input_data("V", df["REB"].to_list())
# TOV
input_data("W", df["TOV"].to_list())
# AST
input_data("X", df["AST"].to_list())
# STL
input_data("Y", df["STL"].to_list())
# BLK
input_data("Z", df["BLK"].to_list())

# 色を塗る
change_color("Tokyo Z")
# ランクをつける
make_rank(target_name="Tokyo Z")
# 書式を保持したまま保存
output_file_name = f'modified_example_{formatted_date}.xlsx'
wb.save(output_file_name)

# メモリ上にエクセルファイルを保存
excel_buffer = BytesIO()
wb.save(excel_buffer)
excel_buffer.seek(0)  # バッファの先頭に戻す

# GCS クライアントを初期化
client = storage.Client()
bucket_name = "b3_output_bucket"  # あなたのバケット名
# バケットと保存先を指定
bucket = client.bucket(bucket_name)
blob = bucket.blob(output_file_name)

# バッファの内容を GCS にアップロード
blob.upload_from_file(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
print(f"{sheet_name} シートを操作して {output_file_name} に保存しました。")

time.sleep(5)

# メールの送信
bucket_name = 'b3_output_bucket'  # GCSバケット名
blob_name = 'modified_example_20241207.xlsx'  # GCS内のファイルパス
sending_mail(bucket_name, blob_name)
